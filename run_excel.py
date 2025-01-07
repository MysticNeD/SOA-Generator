from openpyxl import workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def soa_p2m_template(soa_p2m_sheet, entity_name, entity_code):
    font = Font(bold=True, size = 11, name = 'Arial')

    entity_name_title = soa_p2m_sheet['A1']
    entity_name_title.value = "ENTITY NAME:"
    entity_name_title.font = font

    entity_name_cell = soa_p2m_sheet['C2']
    entity_name_cell.value = entity_name
    entity_name_cell.font = font

    entity_code_title = soa_p2m_sheet['A2']
    entity_code_title.value = "ENTITY CODE:"
    entity_code_title.font = font

    entity_code_cell = soa_p2m_sheet['C1']
    entity_code_cell.value = entity_code
    entity_code_cell.font = font

    supplier_name_title = soa_p2m_sheet['A3']
    supplier_name_title.value = "SUPPLIER NAME:"
    supplier_name_title.font = font
    supplier_name_cell = soa_p2m_sheet['C3']
    supplier_name_cell.value = "BIG COMPANY SDN BHD"
    supplier_name_cell.font = font

    for row in soa_p2m_sheet.iter_rows(min_row=10, max_row=10, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style='thick'), bottom=Side(style='double'))

    for row in soa_p2m_sheet.iter_rows(min_row=13, max_row=13, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    for row in soa_p2m_sheet.iter_rows(min_row=15, max_row=15, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    grand_total = soa_p2m_sheet['C15']
    grand_total.value = "GRAND TOTAL (AS PER SOA)"
    grand_total.font = font
    
    soa_p2m_sheet.merge_cells('A5:I5')
    soa_header = soa_p2m_sheet['A5']
    soa_header.value = "STATEMENT OF ACCOUNT"
    soa_header.font = Font(bold=True, size=11, color='FF0000')
    soa_header.alignment = Alignment(horizontal='center', vertical='center')
    soa_header.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Define the border style
    title_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
    # Apply the border to each cell in the merged range
    for row in soa_p2m_sheet.iter_rows(min_row=5, max_row=5, min_col=1, max_col=8):
        for cell in row:
            cell.border = title_border

    final_title_border = Border(right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Apply the border to each cell in the merged range
    title_border_final = soa_p2m_sheet['I5']
    title_border_final.border = final_title_border
            
    # Column headers
    headers = ['INVOICE \nDATE', 'INVOICE \nNUMBER', 'DESCRIPTION', 'AMOUNT', 'PAID', 'OUTSTANDING','NATURE', 'REMARKS', 'SUPPLIER NAME']
    for col, header in enumerate(headers, start=1):
        cell = soa_p2m_sheet.cell(row=8, column=col)
        cell.value = header
        cell.font = font
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
    
    st_headers = ['STOCK TRF NO.', 'TRF TO', 'TRF FROM', 'SALES INVOICES']
    for col, headers4 in enumerate(st_headers, start=10):
        cell = soa_p2m_sheet.cell(row=11, column = col)
        cell.value = headers4
        cell.font = font
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    down_header = ['As per SOA', 'Netting to be done', 'To Pay', 'Available balance']
    for row, header2 in enumerate(down_header, start=17):
        cell = soa_p2m_sheet.cell(row=row, column=5)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.value = header2
        cell.font = font

    for col in soa_p2m_sheet.iter_rows(min_row = 19, max_row = 19, min_col = 5, max_col = 6):
        for cell in col:
            cell.font = Font(bold=True, size = 11, name = 'Arial', color='FF0000')
            cell.fill = PatternFill(start_color='F2CEEF', end_color='F2CEEF', fill_type='solid')

    for col in soa_p2m_sheet.iter_rows(min_row = 20, max_row = 20, min_col = 6, max_col = 6):
        for cell in col:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    final_header = ['Bank balance as of ', 'To Pay', 'Available balance to retain in bank for backup purpose']
    for row, header3 in enumerate(final_header, start=23):
        cell = soa_p2m_sheet.cell(row=row, column=5)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.value = header3
        cell.font = font

    for col in soa_p2m_sheet.iter_rows(min_row = 24, max_row = 24, min_col = 5, max_col = 6):
        for cell in col:
            cell.font = Font(bold=True, size = 11, name = 'Arial', color='FF0000')
            cell.fill = PatternFill(start_color='F2CEEF', end_color='F2CEEF', fill_type='solid')

    for col in soa_p2m_sheet.iter_rows(min_row = 25, max_row = 25, min_col = 6, max_col = 6):
        for cell in col:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    # Adjust column widths
    column_widths = [10, 12, 35, 13, 12, 20, 15, 15, 35, 17, 8, 11, 18]
    for i, width in enumerate(column_widths, start=1):
        soa_p2m_sheet.column_dimensions[get_column_letter(i)].width = width
    soa_p2m_sheet.row_dimensions[8].height = 28

    print("SOA Sheet Created")

def soa_m2p_template(soa_m2p_sheet, entity_name, entity_code):
    font = Font(bold=True, size = 11, name = 'Arial')

    entity_name_title = soa_m2p_sheet['A1']
    entity_name_title.value = "ENTITY NAME:"
    entity_name_title.font = font

    entity_name_cell = soa_m2p_sheet['C1']
    entity_name_cell.value = "BIG COMPANY SDN BHD"
    entity_name_cell.font = font

    entity_code_title = soa_m2p_sheet['A2']
    entity_code_title.value = "ENTITY CODE:"
    entity_code_title.font = font

    entity_code_cell = soa_m2p_sheet['C2']
    entity_code_cell.value = "3018"
    entity_code_cell.font = font

    supplier_name_title = soa_m2p_sheet['A3']
    supplier_name_title.value = "SUPPLIER NAME:"
    supplier_name_title.font = font
    supplier_name_cell = soa_m2p_sheet['C3']
    supplier_name_cell.value = entity_code
    supplier_name_cell.font = font

    for row in soa_m2p_sheet.iter_rows(min_row=10, max_row=10, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style='thick'), bottom=Side(style='double'))

    for row in soa_m2p_sheet.iter_rows(min_row=13, max_row=13, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    for row in soa_m2p_sheet.iter_rows(min_row=15, max_row=15, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    grand_total = soa_m2p_sheet['C15']
    grand_total.value = "GRAND TOTAL (AS PER SOA)"
    grand_total.font = font
    
    soa_m2p_sheet.merge_cells('A5:I5')
    soa_header = soa_m2p_sheet['A5']
    soa_header.value = "STATEMENT OF ACCOUNT"
    soa_header.font = Font(bold=True, size=11, color='FF0000')
    soa_header.alignment = Alignment(horizontal='center', vertical='center')
    soa_header.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Define the border style
    title_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
    # Apply the border to each cell in the merged range
    for row in soa_m2p_sheet.iter_rows(min_row=5, max_row=5, min_col=1, max_col=8):
        for cell in row:
            cell.border = title_border

    final_title_border = Border(right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Apply the border to each cell in the merged range
    title_border_final = soa_m2p_sheet['I5']
    title_border_final.border = final_title_border
            
    # Column headers
    headers = ['INVOICE \nDATE', 'INVOICE \nNUMBER', 'DESCRIPTION', 'AMOUNT', 'NET OFF', 'OUTSTANDING','NATURE', 'REMARKS', 'SUPPLIER NAME']
    for col, header in enumerate(headers, start=1):
        cell = soa_m2p_sheet.cell(row=8, column=col)
        cell.value = header
        cell.font = font
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
    
    st_headers = ['STOCK TRF NO.', 'TRF TO', 'TRF FROM', 'SALES INVOICES']
    for col, headers4 in enumerate(st_headers, start=10):
        cell = soa_m2p_sheet.cell(row=11, column = col)
        cell.value = headers4
        cell.font = font
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    down_header = ['As per SOA', 'Netting to be done', 'Available balance']
    for row, header2 in enumerate(down_header, start=17):
        cell = soa_m2p_sheet.cell(row=row, column=5)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.value = header2
        cell.font = font

    partial_netting_header = ['Partial Netting Done','', 'Invoice Number', 'Total Amount', 'Previous Netting Amount', 'Current Amount']
    for row, header3 in enumerate(partial_netting_header, start=21):
        cell = soa_m2p_sheet.cell(row=row, column=2)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.value = header3
        cell.font = Font(bold=True, size = 11, name = 'Calibri')
    
    for col in soa_m2p_sheet.iter_rows(min_row = 19, max_row = 19, min_col = 6, max_col = 6):
        for cell in col:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    for col in soa_m2p_sheet.iter_rows(min_row = 26, max_row = 26, min_col = 3, max_col = 3):
        for cell in col:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.border = Border(top=Side(style='thick'), bottom=Side(style='double'))

    # Adjust column widths
    column_widths = [10, 22, 35, 13, 12, 20, 15, 15, 35, 17, 8, 11, 18]
    for i, width in enumerate(column_widths, start=1):
        soa_m2p_sheet.column_dimensions[get_column_letter(i)].width = width
    soa_m2p_sheet.row_dimensions[8].height = 28

    print("SOA Sheet Created")

def p2m_workings(workbook, input_file, output_file):
    arp2m = workbook["AR(P2M)"]
    app2m = workbook["AP(P2M)"]

    # create new sheet and insert title
    workingsp2m = workbook.create_sheet(title="WORKINGS(P2M)")
    headers = ["Invoice_Date", "Invoice No", "Description", "Amount", "Paid", 
            "Outstanding", "Nature", "Remarks", "Supplier Name"]

    for col, header in enumerate(headers, start=4):  # Columns start from 'D' (index 4)
        workingsp2m.cell(row=1, column=col, value=header)

    print("Workings Sheet Created")
    # Filter Open Amount and insert
    header_row = next(app2m.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {header: idx for idx, header in enumerate(header_row)}

    invoice_col_idx = header_map["Invoice Number"]
    open_amount_col_idx = header_map["Open Amount"]



    target_row = 2
    for row in app2m.iter_rows(min_row=1, max_col=99, values_only=True):  
        invoice_no = row[invoice_col_idx]
        open_amount = row[open_amount_col_idx]

        if isinstance(open_amount, (int, float)):
            workingsp2m[f"A{target_row}"] = invoice_no
            workingsp2m[f"G{target_row}"] = open_amount
            target_row += 1

    # text to columns

    for row_index, row in enumerate(workingsp2m.iter_rows(min_col=1, max_col=1), start=1):
        invoice_number = row[0].value

        if invoice_number is not None and "S.T" in invoice_number:
            parts = invoice_number.split(" ")
            if len(parts) > 0:
                workingsp2m.cell(row=row_index, column=1, value=parts[0])  
            if len(parts) > 1:
                workingsp2m.cell(row=row_index, column=2, value=parts[1]) 
            if len(parts) > 2:
                workingsp2m.cell(row=row_index, column=3, value=parts[2]) 
        else:
            workingsp2m.cell(row=row_index, column=1, value=invoice_number)
    print("Text to columns done")
    workingsp2m.cell(row=1, column=1, value="W_RI_Matching")

    # concat RI Number in AR
    ar_max_col = arp2m.max_column
    ar_max_row = arp2m.max_row

    insert_column_index = 1
    arp2m.insert_cols(insert_column_index) 

    header_row = 1
    RI_header = "Doc Type"
    doc_type = None
    for col in range(1, arp2m.max_column + 1):
        if arp2m.cell(row=header_row, column=col).value == RI_header:
            doc_type = col
            break

    DN_header = "Document Number"
    document_no = None
    for col in range(1, arp2m.max_column + 1):
        if arp2m.cell(row = header_row, column = col).value == DN_header:
            document_no = col
            break

    for row_index in range(1, ar_max_row + 1):
        ri_value = arp2m.cell(row = row_index, column = doc_type).value # or column = 4
        number_value = arp2m.cell(row = row_index, column = document_no).value # or column = 5
        ri_number = f"{ri_value}{number_value}"
        arp2m.cell(row = row_index, column = 1, value = ri_number)

    arp2m.cell(row = 1, column = 1, value = "RI_Matching")
    print("RI Matching done, now generating formulas...")
    # remark, invoice no, paid and supplier name
    start_row = 2

    p2m_open_amount_col_idx = None
    p2m_invoice_date_col_idx = None
    p2m_remark_col_idx = None
    headers = []
    for col in range(1, arp2m.max_column + 1):
        header_value = arp2m.cell(row=1, column=col).value
        headers.append(header_value)
        if header_value == "Open Amount":
            p2m_open_amount_col_idx = col
        if header_value == "Invoice Date":
            p2m_invoice_date_col_idx = col
        if header_value == "Remark":
            p2m_remark_col_idx = col

    print(f"Headers found: {headers}")

    if p2m_open_amount_col_idx is None:
        raise ValueError("Header 'Open Amount' not found")
    if p2m_invoice_date_col_idx is None:
        raise ValueError("Header 'Invoice Date' not found")
    if p2m_remark_col_idx is None:
        raise ValueError("Header 'Remark' not found")

    for row in range(2, workingsp2m.max_row + 1):
        formula = f"=VLOOKUP(A{row},'AR(P2M)'!A1:{chr(64 + p2m_invoice_date_col_idx)}{arp2m.max_row},{p2m_invoice_date_col_idx},0)"
        workingsp2m[f"D{row}"] = formula
        cell_date = workingsp2m[f"D{row}"]
        cell_date.number_format = "dd/mm/yyyy"
        print("idate done")

    for row in range(start_row, workingsp2m.max_row + 1):
        formula = f'=A{row}'
        workingsp2m[f"E{row}"] = formula
        print("ino done")

    for row in range(2, workingsp2m.max_row + 1):
        formula = f"=VLOOKUP(A{row},'AR(P2M)'!A1:{chr(64 + p2m_remark_col_idx)}{arp2m.max_row},{p2m_remark_col_idx},0)"
        workingsp2m[f"F{row}"] = formula
        print("desc done")

    for row in range(start_row, workingsp2m.max_row + 1):
        formula = f'=G{row}-I{row}'
        workingsp2m[f"H{row}"] = formula
        print("paid done")

    for row in range(2, workingsp2m.max_row + 1):
        formula = f"=VLOOKUP(A{row},'AR(P2M)'!A1:{chr(64 + p2m_open_amount_col_idx)}{arp2m.max_row},{p2m_open_amount_col_idx},0)"
        workingsp2m[f"I{row}"] = formula
        print("outstanding done")

    for row in range(start_row, workingsp2m.max_row + 1):
        formula = f'=IF(B{row}="S.T","XILNEX: STOCK TRF","")'
        workingsp2m[f"J{row}"] = formula
        print("nature done")

    for row in range(start_row, workingsp2m.max_row + 1):
        formula = f'=IF(B{row}="S.T","REFER TO TAB STOCK TRANSFER","MULTICARE HEALTH PHARMACY SDN")'
        workingsp2m[f"L{row}"] = formula
        print("st done")

def m2p_workings(workbook, input_file, output_file, entity_name, entity_code):
    arm2p = workbook["AR(M2P)"]
    apm2p = workbook["AP(M2P)"]

    # create new sheet and insert title
    workingsm2p = workbook.create_sheet(title="WORKINGS(M2P)")
    headers = ["Invoice_Date", "Invoice No", "Description", "Amount", "Net Off", 
            "Outstanding", "Nature", "Remarks", "Supplier Name"]

    for col, header in enumerate(headers, start=4):  # Columns start from 'D' (index 4)
        workingsm2p.cell(row=1, column=col, value=header)

    print("Workings Sheet Created")
    # Filter Open Amount and insert
    header_row = next(apm2p.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {header: idx for idx, header in enumerate(header_row)}

    invoice_col_idx = header_map["Invoice Number"]
    open_amount_col_idx = header_map["Open Amount"]


    target_row = 2
    for row in apm2p.iter_rows(min_row=1, max_col=99, values_only=True): 
        invoice_no = row[invoice_col_idx]
        open_amount = row[open_amount_col_idx]

        if isinstance(open_amount, (int, float)):
            workingsm2p[f"A{target_row}"] = invoice_no
            workingsm2p[f"G{target_row}"] = open_amount
            target_row += 1

    # text to columns

    for row_index, row in enumerate(workingsm2p.iter_rows(min_col=1, max_col=1), start=1):
        invoice_number = row[0].value

        if invoice_number is not None and "S.T" in invoice_number:
            parts = invoice_number.split(" ")
            if len(parts) > 0:
                workingsm2p.cell(row=row_index, column=1, value=parts[0])  
            if len(parts) > 1:
                workingsm2p.cell(row=row_index, column=2, value=parts[1])
            if len(parts) > 2:
                workingsm2p.cell(row=row_index, column=3, value=parts[2])
        else:
            workingsm2p.cell(row=row_index, column=1, value=invoice_number)
    print("Text to columns done")
    workingsm2p.cell(row=1, column=1, value="W_RI_Matching")

    # concat RI Number in AR
    ar_max_col = arm2p.max_column
    ar_max_row = arm2p.max_row

    insert_column_index = 1
    arm2p.insert_cols(insert_column_index) 

    header_row = 1
    RI_header = "Doc Type"
    doc_type = None
    for col in range(1, arm2p.max_column + 1):
        if arm2p.cell(row=header_row, column=col).value == RI_header:
            doc_type = col
            break

    DN_header = "Document Number"
    document_no = None
    for col in range(1, arm2p.max_column + 1):
        if arm2p.cell(row = header_row, column = col).value == DN_header:
            document_no = col
            break

    for row_index in range(1, ar_max_row + 1):
        ri_value = arm2p.cell(row = row_index, column = doc_type).value # or column = 4
        number_value = arm2p.cell(row = row_index, column = document_no).value # or column = 5
        ri_number = f"{ri_value}{number_value}"
        arm2p.cell(row = row_index, column = 1, value = ri_number)

    arm2p.cell(row = 1, column = 1, value = "RI_Matching")
    print("RI Matching done, now generating formulas...")
    # remark, invoice no, paid and supplier name
    start_row = 2

    m2p_open_amount_col_idx = None
    m2p_invoice_date_col_idx = None
    m2p_remark_col_idx = None
    headers = []
    for col in range(1, arm2p.max_column + 1):
        header_value = arm2p.cell(row=1, column=col).value
        headers.append(header_value)
        if header_value == "Open Amount":
            m2p_open_amount_col_idx = col
        if header_value == "Invoice Date":
            m2p_invoice_date_col_idx = col
        if header_value == "Remark":
            m2p_remark_col_idx = col

    if m2p_open_amount_col_idx is None:
        raise ValueError("Header 'Open Amount' not found")
    if m2p_invoice_date_col_idx is None:
        raise ValueError("Header 'Invoice Date' not found")
    if m2p_remark_col_idx is None:
        raise ValueError("Header 'Remarks' not found")


    for row in range(2, workingsm2p.max_row + 1):
        formula = f"=VLOOKUP(A{row},'AR(M2P)'!A1:{chr(64 + m2p_invoice_date_col_idx)}{arm2p.max_row},{m2p_invoice_date_col_idx},0)"
        workingsm2p[f"D{row}"] = formula
        cell_date = workingsm2p[f"D{row}"]
        cell_date.number_format = "dd/mm/yyyy"
        print("M2P idate done")

    for row in range(start_row, workingsm2p.max_row + 1):
        formula = f'=A{row}'
        workingsm2p[f"E{row}"] = formula
        print("M2P ino done")

    for row in range(2, workingsm2p.max_row + 1):
        formula = f"=VLOOKUP(A{row},'AR(M2P)'!A1:{chr(64 + m2p_remark_col_idx)}{arm2p.max_row},{m2p_remark_col_idx},0)"
        workingsm2p[f"F{row}"] = formula
        print("M2P desc done")

    for row in range(start_row, workingsm2p.max_row + 1):
        formula = f'=G{row}-I{row}'
        workingsm2p[f"H{row}"] = formula
        print("M2P paid done")

    for row in range(2, workingsm2p.max_row + 1):
        formula = f"=VLOOKUP(A{row},'AR(M2P)'!A1:{chr(64 + m2p_open_amount_col_idx)}{arm2p.max_row},{m2p_open_amount_col_idx},0)"
        workingsm2p[f"I{row}"] = formula
        print("M2P outstanding done")

    for row in range(start_row, workingsm2p.max_row + 1):
        formula = f'=IF(B{row}="S.T","XILNEX: STOCK TRF","")'
        workingsm2p[f"J{row}"] = formula
        print("M2P nature done")

    for row in range(start_row, workingsm2p.max_row + 1):
        formula = f'=IF(B{row}="S.T","REFER TO TAB STOCK TRANSFER","{entity_code}")'
        workingsm2p[f"L{row}"] = formula
        print("M2P st done")

def generate_soa(input_file, output_file, entity_name, entity_code):  
    print("Generating starts now... Please do not close the program.")
    workbook = load_workbook(input_file)

    soa_p2m_sheet = workbook.create_sheet(title="SOA(P2M)")
    soa_p2m_template(soa_p2m_sheet, entity_name, entity_code)
    soa_m2p_sheet = workbook.create_sheet(title="SOA(M2P)")
    soa_m2p_template(soa_m2p_sheet, entity_name, entity_code)

    p2m_workings(workbook, input_file, output_file)
    m2p_workings(workbook, input_file, output_file, entity_name, entity_code)

    print("Generate done. Now saving file...")
    workbook.save(output_file)
    print(f"Processing complete. File saved as '{output_file}'.")

